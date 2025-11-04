from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging
class ThinkExcel:
    def __init__(self):
        self.logger = logging.getLogger("tsfbatch")
        self.build_d_alpha()
    def open_wb(self,wb,ws=None):
        self.wb = load_workbook(wb)
        self.current_wb = wb
        if ws:
            self.ws = self.wb.get_worksheet_by_name(ws)
        else:
            self.ws = self.wb.active
    def activate_ws(self,ws):
        for i, s in enumerate(self.wb.sheetnames):
            if s == ws:
                break
        self.wb.active = i
        self.ws = self.wb.active
    def create_wb(self,file):
        self.wb = Workbook()
        self.wb_name = file
        # the active worksheet
        # self.ws = self.wb.active
        # self.current_wb = file
        # self.wb.save(file)
    def create_ws(self,ws):
        self.wb.create_sheet(ws)
        self.ws = self.activate_ws(ws)

    def write_to_cw(self,**kwargs):
        _d = {}
        self.ws = self.wb.active
        _d.update(kwargs)
        for k,v in _d.items():
            self.ws[k] = v
    def write_to_cell_range(self,row,col,val):
        self.ws.cell(row=row,column=col).value = val
    def save_current_wb(self):
        try:
            tsheet=self.wb.get_sheet_by_name('Sheet')
            if tsheet:
                self.wb.remove_sheet(tsheet)
        except Exception:
            pass
        self.wb.save(self.wb_name)
    def set_style(self,cell):
        _cell = self.ws[cell]
        _cell.font = Font(name="Tahoma", size=16, color="00339966")
    def format_header(self,cell):
        font = Font(name='Calibri',
                         size=12,
                         bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')
        fill = PatternFill(fill_type=None, start_color='FFFC33',end_color='FFFC33')
        _cell = self.ws[cell]
        _cell.font = font
        _cell.fill = PatternFill("solid", fgColor="DDDDDD")

    def add_df_to_ws(self,ws,new=False,df=None):
        if new:
            self.create_ws(ws)
            self.activate_ws(ws)
        else:
            self.activate_ws(ws)
        for r in dataframe_to_rows(df, index=True, header=True):
            self.ws.append(r)
    def add_table_to_ws(self,dispTabName,ref):
        tab = Table(displayName=dispTabName, ref=ref)
        style = TableStyleInfo(name="TableStyleDark4", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        self.ws.add_table(tab)
        self.auto_fit_col_width(self.ws)
    def auto_fit_col_width(self,ws):
        try:
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = (value+2)
        except Exception as e:
            print(e)

    def build_d_alpha(self):         
        alphabets = [ch for ch in map(chr,range(65,91))]
        alphabets_extended = [f"{char1}{char0}" for char0 in alphabets  for char1 in alphabets[:10]]
        alphabets_extended = alphabets + alphabets_extended
        self.d_alpha = {indx:alpha for indx,alpha in enumerate(alphabets_extended)} 
    
    def get_ref(self,start_row,start_col,end_row,end_col):
        return  f"{self.d_alpha[start_row-1]}{start_col}:{self.d_alpha[end_col-1]}{end_row}"



if __name__=='__main__':
    tx = ThinkExcel()
    # tx.create_wb("my_first_python_excel.xlsx")
    # d = {"A1":100,"B1":200,"C1":400,"A2":"My first Excel coding in python"}
    # tx.write_to_cw(**d)
    # tx.write_to_cell_range(row=3,col=1,val="This written in cell range")
    # tx.save_current_wb()
    tx.open_wb("my_first_python_excel.xlsx")
    tx.write_to_cell_range(row=4,col=1,val='Something written in the saved file')

    tx.create_ws("My second sheet")
    tx.set_style('A1')
    d = {"A1":"This is in the second ws i have just created"}
    tx.write_to_cw(**d)
    tx.save_current_wb()
